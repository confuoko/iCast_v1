# experiment1.py
from openpyxl import load_workbook


def read_excel_with_merged_and_styles(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    print("📘 Объединённые диапазоны:")
    for merged_range in sheet.merged_cells.ranges:
        print("  →", merged_range)

    print("\n📄 Данные с учётом объединений и стилей:\n")

    merged_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = top_left_value

    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            val = cell.value or merged_map.get((cell.row, cell.column))

            font = cell.font
            fill = cell.fill
            alignment = cell.alignment

            print(
                f"({cell.coordinate}) "
                f"'{val}' | font={font.name}, bold={font.bold}, size={font.size}, "
                f"align={alignment.horizontal}, fill={fill.fgColor.rgb}"
            )
        print("-" * 80)


if __name__ == "__main__":
    read_excel_with_merged_and_styles("score_v1.xlsx")
