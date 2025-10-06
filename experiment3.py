# experiment1.py
from openpyxl import load_workbook

def read_excel_clean(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    print("📘 Объединённые диапазоны:")
    for merged_range in sheet.merged_cells.ranges:
        print("  →", merged_range)

    print("\n📄 Данные с учётом объединений (столбцы A–C, без первых 3 строк):\n")

    # === создаём карту объединённых ячеек ===
    merged_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = top_left_value

    # === читаем только нужные данные ===
    all_rows = []
    for row in sheet.iter_rows(min_row=4,  # пропускаем первые 3 строки
                               min_col=1,  # столбец A
                               max_col=3,  # столбец C
                               values_only=False):
        values = [
            (cell.value or merged_map.get((cell.row, cell.column)))
            for cell in row
        ]

        # Пропускаем полностью пустые строки
        if not any(values):
            continue

        all_rows.append(values)

    # === печатаем результат ===
    for row in all_rows:
        print(tuple(row))


if __name__ == "__main__":
    read_excel_clean("score_v2.xlsx")
