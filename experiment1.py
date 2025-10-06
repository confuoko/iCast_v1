# experiment1.py
from openpyxl import load_workbook

def read_excel_with_merged(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    print("📘 Объединённые диапазоны:")
    for merged_range in sheet.merged_cells.ranges:
        print("  →", merged_range)

    print("\n📄 Данные с учётом объединений:\n")
    merged_map = {}

    # Создаём карту объединений: каждая ячейка знает, из какого диапазона она
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = top_left_value

    # Проходим по всем строкам и выводим данные, подставляя значение из объединений
    for row in sheet.iter_rows(values_only=False):
        values = []
        for cell in row:
            val = cell.value or merged_map.get((cell.row, cell.column))
            values.append(val)
        print(tuple(values))

if __name__ == "__main__":
    read_excel_with_merged("score_v2.xlsx")
